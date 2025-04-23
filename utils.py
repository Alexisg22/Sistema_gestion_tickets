from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

#Clase para dar formato al archivo .xlsx descargado 
class ExcelFormatter:
    @staticmethod
    def apply_format(archivo):
        # Cargar el libro de trabajo existente
        wb = load_workbook(archivo)
        ws = wb.active
        
        # Aplicar estilos de encabezado
        ExcelFormatter._format_headers(ws)
        
        # Ajustar ancho de columnas
        ExcelFormatter._adjust_column_width(ws)
        
        # Formatear contenido de celdas
        ExcelFormatter._format_cell_content(ws)
        
        # Aplicar filtro y congelar paneles
        ExcelFormatter._apply_filter_and_freeze(ws)
        
        # Guardar el libro con formato
        wb.save(archivo)
    
    @staticmethod
    def _format_headers(ws):
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
    
    @staticmethod
    def _adjust_column_width(ws):
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20
    
    @staticmethod
    def _format_cell_content(ws):
        data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Colores para el semáforo
        verde_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        naranja_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        rojo_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        
        # Ubicar columnas especiales
        semaforo_col = None
        estado_col = None
        
        for col in range(1, ws.max_column + 1):
            header_value = ws.cell(row=1, column=col).value
            if header_value == "Semáforo":
                semaforo_col = col
            elif header_value == "Estado":
                estado_col = col
        
        # Formatear celdas
        for row in range(2, ws.max_row + 1):
            # Aplicar bordes y alineación a todas las celdas
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = data_alignment
                cell.border = thin_border
            
            # Aplicar color según valor de semáforo
            if semaforo_col:
                cell = ws.cell(row=row, column=semaforo_col)
                valor = cell.value
                if valor == "verde":
                    cell.fill = verde_fill
                elif valor == "naranja":
                    cell.fill = naranja_fill
                elif valor == "rojo":
                    cell.fill = rojo_fill
            
            # Aplicar formato a la columna de estado
            if estado_col:
                cell = ws.cell(row=row, column=estado_col)
                if cell.value and "cerrado" in str(cell.value).lower():
                    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    @staticmethod
    def _apply_filter_and_freeze(ws):
        # Aplicar filtro a los encabezados
        ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=ws.max_column).column_letter}1"
        
        # Congelar la primera fila
        ws.freeze_panes = "A2"

